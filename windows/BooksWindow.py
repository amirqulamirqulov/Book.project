from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import PyQt5
import traceback


from models import Book_type, Book_name, Book
from openpyxl import Workbook


class BookWindow(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.initUI()


    def initUI(self):
    
        self.resize(800, 450)

        self.qlb_book_type = QLabel("Book type", self)
        self.qlb_book_type.move(40, 30)

        self.cbb_book_type = QComboBox(self)
        self.cbb_book_type.move(100, 30) 

        self.qlb_book_name = QLabel("Book name", self)
        self.qlb_book_name.move(170, 30)

        self.cbb_book_name = QComboBox(self)
        self.cbb_book_name.move(230, 30) 
        self.cbb_book_name.setGeometry(230, 30, 108, 23)

        self.qlb_book_year = QLabel("Book year", self)
        self.qlb_book_year.move(40, 80)

        self.le_book_year = QLineEdit(self)
        self.le_book_year.move(150, 80)
        self.le_book_year.setGeometry(150, 80, 185, 20)

        self.qlb_book_countpage = QLabel("Book count page", self)
        self.qlb_book_countpage.move(40, 120)

        self.le_book_countpage = QLineEdit(self)
        self.le_book_countpage.move(150, 120)
        self.le_book_countpage.setGeometry(150, 120, 185, 20)

        self.qlb_book_price = QLabel("Book price", self)
        self.qlb_book_price.move(40, 160)

        self.le_book_price = QLineEdit(self)
        self.le_book_price.move(150, 160)
        self.le_book_price.setGeometry(150, 160, 185, 20)

        self.qlb_book_author = QLabel("Book author", self)
        self.qlb_book_author.move(40, 200)

        self.le_book_author = QLineEdit(self)
        self.le_book_author.move(150, 200)
        self.le_book_author.setGeometry(150, 200, 185, 20)

        self.table = QTableWidget(self)
        self.table.setGeometry(370, 30, 900, 620)
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            ["book type id", 'book_type name', 'book name id', 'books name',
            "book id", "book year", "book count_page", "book price", "book author"])
        self.table.setRowCount(0)
        self.table.hideColumn(0)
        self.table.hideColumn(2)
        self.table.hideColumn(4)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setColumnWidth(3, 200)
        self.table.setColumnWidth(8, 275)


        self.btn_add = QPushButton("Add", self)
        self.btn_add.setFont((QFont('Times New Roman',18)))
        self.btn_add.setGeometry(40, 260, 295, 60)
        self.btn_add.clicked.connect(self.onAdd)

        self.btn_upd = QPushButton("Update", self)
        self.btn_upd.setFont((QFont('Times New Roman',18)))
        self.btn_upd.setGeometry(40, 370, 295, 60)
        self.btn_upd.clicked.connect(self.onUpd)

        self.btn_del = QPushButton("Delete", self)
        self.btn_del.setFont((QFont('Times New Roman',18)))
        self.btn_del.setGeometry(40, 480, 295, 60)
        self.btn_del.clicked.connect(self.onDel)

        self.btn_rep = QPushButton("Report", self)
        self.btn_rep.setFont((QFont('Times New Roman',18)))
        self.btn_rep.setGeometry(40, 590, 295, 60)
        self.btn_rep.clicked.connect(self.onRep)

        for book in Book.objects():
            book_name = book.Book_name
            book_type = book_name.Book_type
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(book_type.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(str(book_type.Name)))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(str(book_name.id)))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(book_name.Name)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(book.id)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(book.Year)))
            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(book.Count_page)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(book.Price)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(str(book.Author)))
        
        self.table.clicked.connect(self.onClicked)

        self.cbb_book_type.currentIndexChanged.connect(self.changedBT)
        for book in Book_type.objects():
            self.cbb_book_type.addItem(book.Name, book.id)
        self.cbb_book_name.currentIndexChanged.connect(self.fillTable)


    def onRep(self):
        wb = Workbook()
        try:
            # grab the active worksheet
            ws = wb.active
            ws[f'A{1}'] = "Book type"
            ws[f'B{1}'] = "Book name"
            ws[f'C{1}'] = "Book year"
            ws[f'D{1}'] = "Book count page"
            ws[f'E{1}'] = "Book price"
            ws[f'F{1}'] = "Book author"

            for sel_row in range(self.table.rowCount()):
                type = self.table.item(sel_row, 1).text()
                name = self.table.item(sel_row, 3).text()
                year = int(self.table.item(sel_row, 5).text())
                countpage = int(self.table.item(sel_row, 6).text())
                price = int(self.table.item(sel_row, 7).text())
                author = self.table.item(sel_row, 8).text()

                ws[f'A{sel_row + 2}'] = type
                ws[f'B{sel_row + 2}'] = name
                ws[f'C{sel_row + 2}'] = year
                ws[f'D{sel_row + 2}'] = countpage
                ws[f'E{sel_row + 2}'] = price
                ws[f'F{sel_row + 2}'] = author
            # Save the file
            wb.save("books.xlsx")
            wb.close()

        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()
            wb.close()
            traceback.print_exc()
        
    
    def changedBT(self):
        bk_id = self.cbb_book_type.currentData()
        self.cbb_book_name.clear()
        for book in Book_name.objects():
            if book.Bookid == bk_id:
                self.cbb_book_name.addItem(book.Name, book.id)

    def onClicked(self):
        self.sel_row = self.table.currentRow()
        self.le_book_year.setText(self.table.item(self.sel_row, 5).text())
        self.le_book_countpage.setText(self.table.item(self.sel_row, 6).text())
        self.le_book_price.setText(self.table.item(self.sel_row, 7).text())
        self.le_book_author.setText(self.table.item(self.sel_row, 8).text())

        self.sel_Book = Book(int(self.table.item(
            self.sel_row, 5).text()), int(self.table.item(self.sel_row, 6).text())
            , int(self.table.item(self.sel_row, 7).text()), self.table.item(self.sel_row, 8).text(),
            self.table.item(self.sel_row, 4).text(),self.table.item(self.sel_row, 0).text())

    
    def onAdd(self):
        book_type = Book_type.get_by_id(self.cbb_book_type.currentData())
        book_name = Book_name.get_by_id(self.cbb_book_name.currentData())

        book = Book(int(self.le_book_year.text()), int(self.le_book_countpage.text()),
        int(self.le_book_price.text()), self.le_book_author.text(), book_name.id)
        book.save()

        row_count = self.table.rowCount()
        self.table.setRowCount(row_count + 1)
        self.table.setItem(row_count, 0,
                           QTableWidgetItem(str(book_type.id)))
        self.table.setItem(row_count, 1,
                           QTableWidgetItem(str(book_type.Name)))
        self.table.setItem(row_count, 2,
                           QTableWidgetItem(str(book_name.id)))
        self.table.setItem(row_count, 3,
                           QTableWidgetItem(str(book_name.Name)))
        self.table.setItem(row_count, 4,
                           QTableWidgetItem(str(book.id)))
        self.table.setItem(row_count, 5,
                           QTableWidgetItem(str(book.Year)))
        self.table.setItem(row_count, 6,
                           QTableWidgetItem(str(book.Count_page)))
        self.table.setItem(row_count, 7,
                           QTableWidgetItem(str(book.Price)))
        self.table.setItem(row_count, 8,
                           QTableWidgetItem(str(book.Author)))

    def onUpd(self):
        if  self.sel_Book  is not None:
            
            book_year = int(self.le_book_year.text())
            book_countpage = int(self.le_book_countpage.text())
            book_price = int(self.le_book_price.text())
            book_author = self.le_book_author.text()
            book_id = self.cbb_book_name.currentData()
            book_name_id = int(self.table.item(self.sel_row, 4).text())

            bookk = Book(book_year, book_countpage, book_price,
            book_author, book_id, book_name_id)
            bookk.save()

            self.fillTable()

    def onDel(self):
        if  self.sel_Book  is not None:
            
            book_year = int(self.le_book_year.text())
            book_countpage = int(self.le_book_countpage.text())
            book_price = int(self.le_book_price.text())
            book_author = self.le_book_author.text()
            book_id = self.cbb_book_name.currentData()
            book_name_id = int(self.table.item(self.sel_row, 4).text())

            bookk = Book(book_year, book_countpage, book_price,
            book_author, book_id, book_name_id)
            bookk.delete()

            self.fillTable()

    def fillTable(self):
        self.table.clear()
        self.table.setHorizontalHeaderLabels(
            ["Book_type id", "Book type name", "Bookid", "Book name",
             "Bookid", "Book year", "Book count page", "Book price"
             , "Book author"])
        self.table.setRowCount(0)
        self.table.hideColumn(0)
        self.table.hideColumn(2)
        self.table.hideColumn(4)
        current_id = self.cbb_book_name.currentData()
        for book in Book.objects():
            if current_id == book.Bookid:
                book_name = book.Book_name
                book_type = book_name.Book_type
                rowCount = self.table.rowCount()
                self.table.setRowCount(rowCount + 1)
                self.table.setItem(rowCount, 0,
                                   QTableWidgetItem(str(book_type.id)))
                self.table.setItem(rowCount, 1,
                                   QTableWidgetItem(str(book_type.Name)))
                self.table.setItem(rowCount, 2,
                                   QTableWidgetItem(str(book_name.id)))
                self.table.setItem(rowCount, 3,
                                   QTableWidgetItem(str(book_name.Name)))
                self.table.setItem(rowCount, 4,
                                   QTableWidgetItem(str(book.id)))
                self.table.setItem(rowCount, 5,
                                   QTableWidgetItem(str(book.Year)))
                self.table.setItem(rowCount, 6,
                                   QTableWidgetItem(str(book.Count_page)))
                self.table.setItem(rowCount, 7,
                                   QTableWidgetItem(str(book.Price)))
                self.table.setItem(rowCount, 8,
                                   QTableWidgetItem(str(book.Author)))
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(3, 200)
        self.table.setColumnWidth(8, 275)


